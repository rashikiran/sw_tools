#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Alignment

def merge_first_row_in_all_sheets(excel_file):
    try:
        # Load the existing workbook
        wb = openpyxl.load_workbook(excel_file)

        # Loop through all sheets in the workbook
        for sheet in wb.sheetnames:
            worksheet = wb[sheet]
            
            # Get the number of columns in the first row
            max_col = worksheet.max_column
            
            # Merge cells in the first row across all columns
            worksheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=max_col)
            
            # Set the merged cell's value to the sheet name or any desired header text
            worksheet.cell(row=3, column=3).value = sheet  # Using sheet name as the merged header

            # Optionally, center the text in the merged cell
            worksheet.cell(row=3, column=3).alignment = Alignment(horizontal='center', vertical='center')
        
        # Save the workbook after modifications
        wb.save(excel_file)
        print(f"First row in all sheets merged successfully in {excel_file}")
    
    except Exception as e:
        print(f"Error processing the file: {e}")

# Example usage
excel_file = 'contexts.xlsx'  # Replace with your Excel file
merge_first_row_in_all_sheets(excel_file)
