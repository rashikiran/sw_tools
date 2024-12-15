#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Alignment, PatternFill
import sys

def merge_first_row_in_all_sheets(excel_file):
    try:
        # Load the existing workbook
        wb = openpyxl.load_workbook(excel_file)

        # Define the light grey background color fill
        light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="lightDown")

        # Loop through all sheets in the workbook
        for sheet in wb.sheetnames:
            worksheet = wb[sheet]

            # Get the number of columns in the first row
            max_col = worksheet.max_column

            # Merge cells in the first row across all columns
            worksheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=max_col)

            # Set the merged cell's value to the sheet name or any desired header text
            worksheet.cell(row=3, column=3).value = sheet  # Using sheet name as the merged header

            # Optionally, center the text in the merged cell
            worksheet.cell(row=3, column=3).alignment = Alignment(horizontal='center', vertical='center')

            # Fill the merged cells with light grey background
            for col in range(3, max_col + 1):  # Columns C to max_col
                worksheet.cell(row=3, column=col).fill = light_grey_fill

        # Save the workbook after modifications
        wb.save(excel_file)
        print(f"First row in all sheets merged and filled with light grey color successfully in {excel_file}")

    except Exception as e:
        print(f"Error processing the file: {e}")

# Check if an output Excel file name was provided via command line argument
if len(sys.argv) < 2:
    print("Please provide the output Excel file name as a command line argument.")
    sys.exit(1)

# Get the output Excel file name from the first command line argument
excel_file = sys.argv[1]

merge_first_row_in_all_sheets(excel_file)

