#!/usr/bin/env python3

import openpyxl
import sys
from openpyxl.styles import Alignment, PatternFill

def merge_cells_in_range(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Define the sky blue color fill
    sky_blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Iterate over all rows in the sheet
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
            # Check if the value in column D (index 3) is 'Header'
            if row[3].value == "Header":  # Column D is index 3 (0-based index)
                # Merge cells from C to I in the current row
                sheet.merge_cells(
                    start_row=row[2].row, start_column=3,  # Column C = 3
                    end_row=row[2].row, end_column=9      # Column I = 9
                )
                print(f"Merged cells C{row[2].row} to I{row[2].row} in sheet {sheet_name}.")

                # Center-align the merged cells
                merged_cell = sheet.cell(row=row[2].row, column=3)  # Reference the starting cell (C)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                print(f"Center-aligned merged cells C{row[2].row} to I{row[2].row} in sheet {sheet_name}.")

                # Fill the merged cells with sky blue background
                for col in range(3, 10):  # Columns C to I (3 to 9)
                    sheet.cell(row=row[2].row, column=col).fill = sky_blue_fill
                print(f"Filled merged cells C{row[2].row} to I{row[2].row} with sky blue color in sheet {sheet_name}.")

    # Save the modified workbook with the same name, but with 'modified_' prefix
    output_file = file_path
    workbook.save(output_file)
    print(f"Modified file saved as {output_file}")

# Main function to accept command line arguments
def main():
    if len(sys.argv) != 2:
        print("Usage: python script.py <input_excel_file>")
        sys.exit(1)

    # Get input file from command line argument
    file_path = sys.argv[1]

    # Call function to merge cells, center align, and fill with color
    merge_cells_in_range(file_path)

# Run the main function when the script is executed
if __name__ == "__main__":
    main()

